from typing import Any, Dict, List, Optional
from fastapi import HTTPException, Request, UploadFile, status
from app.modules.auth.services.authenticated.authenticated_service import AuthenticatedService, CustomJSONResponseException
from app.modules.core.models.field_translation_keys import DEFAULT_LANGUAGE
from app.modules.core.models.mapping_keys import CollectionKey
from app.modules.core.services.debug.debug_service import DebugService
from app.modules.core.services.response.response_service import ResponseService
from app.modules.core.services.converter.converter_service import ConverterService
from app.modules.core.services.model.model_service import ModelService
from app.modules.core.enums.type_enum import EGender, OutputDataType
from app.modules.core.types.response import CustomJSONResponse


class AgentBulkController(
    DebugService,
    ResponseService,
    ConverterService,
    AuthenticatedService,
    ModelService,
):
    def __init__(self, accept_language: Optional[str] = DEFAULT_LANGUAGE):
        from app.modules.core.services.generic.generic_services import GenericService

        self.accept_language = accept_language
        self.generic_service = GenericService()

    GENDER_LABEL_MAP = {
        "m": EGender.MALE.value,
        "male": EGender.MALE.value,
        "masculin": EGender.MALE.value,
        "homme": EGender.MALE.value,
        "mobali": EGender.MALE.value,
        "f": EGender.FEMALE.value,
        "female": EGender.FEMALE.value,
        "féminin": EGender.FEMALE.value,
        "feminin": EGender.FEMALE.value,
        "femme": EGender.FEMALE.value,
        "mwasi": EGender.FEMALE.value,
    }

    def _resolve_gender(self, raw: str) -> Optional[str]:
        if not raw:
            return None
        return self.GENDER_LABEL_MAP.get(raw.strip().lower())

    async def _resolve_ref_by_name(self, collection_key, name: str, user_details=None) -> Optional[str]:
        """Resolve a cfg_function or cfg_grade by name (case-insensitive)."""
        if not name:
            return None
        item = await self.generic_service.fetch_one_from_collection(
            collection_key=collection_key,
            output_data_type=OutputDataType.DEFAULT.value,
            query={"filter__name": name},
            user=user_details,
        )
        if item:
            return str(item.get('_id', ''))
        return None

    # ── Bulk Upload ──────────────────────────────────────────────────

    async def bulk_upload_agents(self, request: Request, file: UploadFile, cfg_organism_chart_id: Optional[str] = None):
        """
        Parse an Excel file and bulk-create organization agent records.
        Expected columns:
          A: matricule (required)
          B: first_name (required)
          C: last_name (required)
          D: sur_name (optional)
          E: gender (required — m/f)
          F: fonction (optional — name of cfg_function)
          G: grade (optional — name of cfg_grade)
        """
        try:
            import openpyxl
            from io import BytesIO

            user_details = await self.get_user_info(request, self.accept_language)
            sys_organization_id = user_details.get('sys_organization_id')
            await self.get_api_consumer(request, self.accept_language)
            await self.get_user_profil(request, self.accept_language)
            self.app_debug_print(f"\n\n\n user_details : {user_details['id']}",True)
            contents = await file.read()
            wb = openpyxl.load_workbook(BytesIO(contents), read_only=True)
            ws = wb.active

            created_count = 0
            errors: List[Dict[str, Any]] = []
            row_num = 0
            self.app_debug_print(f"\n\n\n ws : {ws}",True)
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_num += 1
                if not row or not row[0]:
                    self.app_debug_print(f"\n\n\n continue : not row",True)
                    continue

                matricule = str(row[0]).strip() if row[0] else ''
                first_name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                last_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                sur_name = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                raw_gender = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                raw_function = str(row[5]).strip() if len(row) > 5 and row[5] else ''
                raw_grade = str(row[6]).strip() if len(row) > 6 and row[6] else ''

                if not matricule:
                    self.app_debug_print(f"\n\n\n continue : not matricule",True)
                    errors.append({"row": row_num + 1, "error": "Matricule manquant"})
                    continue

                if not first_name:
                    self.app_debug_print(f"\n\n\n continue : not first name",True)
                    errors.append({"row": row_num + 1, "error": "Prénom manquant"})
                    continue

                if not last_name:
                    self.app_debug_print(f"\n\n\n continue : not last name",True)
                    errors.append({"row": row_num + 1, "error": "Nom manquant"})
                    continue

                gender = self._resolve_gender(raw_gender)
                if not gender:
                    self.app_debug_print(f"\n\n\n continue : not gender",True)
                    errors.append({"row": row_num + 1, "error": f"Genre invalide: '{raw_gender}'. Valeurs acceptées: m, f"})
                    continue

                # Check if agent with this matricule already exists
                existing = await self.generic_service.fetch_one_from_collection(
                    collection_key=CollectionKey.SYS_ORGANIZATION_AGENT,
                    output_data_type=OutputDataType.DEFAULT.value,
                    query={"filter__matricule": matricule},
                    user=user_details,
                )
                if existing:
                    self.app_debug_print(f"\n\n\n continue : existing : {existing}",True)
                    errors.append({"row": row_num + 1, "error": f"Agent avec matricule '{matricule}' existe déjà"})
                    continue

                # Resolve function and grade by name (optional)
                self.app_debug_print(f"\n\n\n continue : raw_function : {raw_function}",True)
                cfg_function_id = None
                if raw_function:
                    cfg_function_id = await self._resolve_ref_by_name(CollectionKey.CFG_FUNCTION, raw_function, user_details=user_details)
                    if not cfg_function_id:
                        self.app_debug_print(f"\n\n\n continue : cfg_function_id : {cfg_function_id}",True)
                        errors.append({"row": row_num + 1, "error": f"Fonction introuvable: '{raw_function}'"})
                        # continue

                cfg_grade_id = None
                self.app_debug_print(f"\n\n\n continue : raw_grade : {raw_grade}",True)
                if raw_grade:
                    cfg_grade_id = await self._resolve_ref_by_name(CollectionKey.CFG_GRADE, raw_grade, user_details=user_details)
                    if not cfg_grade_id:
                        self.app_debug_print(f"\n\n\n continue : cfg_grade_id : {cfg_grade_id}",True)
                        errors.append({"row": row_num + 1, "error": f"Grade introuvable: '{raw_grade}'"})
                        # continue

                try:
                    # Step 1: Create person
                    person_data = {
                        "first_name": first_name,
                        "last_name": last_name,
                        "gender": gender,
                        "sys_organization_id": sys_organization_id,
                    }
                    if sur_name:
                        person_data["sur_name"] = sur_name
                    self.app_debug_print(f"\n\n\n person_data : {person_data}",True)
                    person_saved_id = await self.generic_service.add_data_to_collection(
                        CollectionKey.SYS_PERSON, person_data,
                        user=user_details, request=request
                    )

                    self.app_debug_print(f"\n\n\n person_saved_id : {person_saved_id}",True)

                    # Step 2: Create agent
                    agent_data: Dict[str, Any] = {
                        "matricule": matricule,
                        "sys_person_id": person_saved_id,
                        "sys_organization_id": sys_organization_id,
                    }
                    if cfg_function_id:
                        agent_data["cfg_function_id"] = cfg_function_id
                    if cfg_grade_id:
                        agent_data["cfg_grade_id"] = cfg_grade_id
                    if cfg_organism_chart_id:
                        agent_data["cfg_organism_chart_id"] = cfg_organism_chart_id
                    self.app_debug_print(f"\n\n\n agent_data : {agent_data}",True)
                    await self.generic_service.add_data_to_collection(
                        CollectionKey.SYS_ORGANIZATION_AGENT, agent_data,
                        user=user_details, request=request
                    )
                    created_count += 1
                except Exception as row_err:
                    errors.append({"row": row_num + 1, "error": str(row_err)})

            wb.close()

            return CustomJSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "status_code": status.HTTP_200_OK,
                    "message": f"Import terminé. Créés: {created_count}, Erreurs: {len(errors)}",
                    "data": {
                        "created_count": created_count,
                        "error_count": len(errors),
                        "errors": errors[:50],
                    },
                },
            )

        except PermissionError as e:
            raise HTTPException(status_code=403, detail=str(e))
        except HTTPException as e:
            raise e
        except CustomJSONResponseException:
            raise
        except Exception as e:
            self.app_debug_print(f"Error in bulk_upload_agents: {str(e)}", True)
            raise HTTPException(status_code=500, detail=f"Bulk upload failed: {str(e)}")

    # ── Template Download ────────────────────────────────────────────

    async def download_bulk_template(self, request: Request):
        """Generate and return an Excel template for bulk agent creation."""
        try:
            import openpyxl
            from io import BytesIO
            from fastapi.responses import StreamingResponse

            await self.get_user_info(request, self.accept_language)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Agents"

            headers = [
                'matricule', 'first_name', 'last_name',
                'sur_name', 'gender', 'fonction', 'grade',
            ]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="057aeb", end_color="057aeb", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            col_widths = [18, 20, 20, 18, 12, 22, 22]
            for i, w in enumerate(col_widths, start=1):
                ws.column_dimensions[chr(64 + i)].width = w

            examples = [
                ("MAT-001", "Jean", "Kabongo", "Pierre", "m", "Conducteur", "Agent"),
                ("MAT-002", "Marie", "Lukusa", "", "f", "Percepteur", "Chef d'équipe"),
                ("MAT-003", "Patrick", "Mbuyi", "Joseph", "m", "", ""),
            ]
            example_font = Font(color="999999", italic=True)
            for i, row_data in enumerate(examples, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=i, column=col_idx, value=value)
                    cell.font = example_font

            output = BytesIO()
            wb.save(output)
            wb.close()
            output.seek(0)

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={
                    "Content-Disposition": "attachment; filename=template_agents.xlsx",
                },
            )

        except Exception as e:
            self.app_debug_print(f"Error in download_bulk_template: {str(e)}", True)
            raise HTTPException(status_code=500, detail=f"Template generation failed: {str(e)}")
